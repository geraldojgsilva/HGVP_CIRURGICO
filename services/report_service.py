from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

COLORS = {
    0: "FFECB58C",
    1: "FFFCF6EE"
}

def generate_automation_items_excel(data) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consultas"
    ws.sheet_properties.defaultRowHeight = 20

    columns = [
        ("Nome Paciente", "name", 50),
        ("Numero", "number", 25),
        ("Data Agenda", "date", 25),
        ("Data Envio Mensagem", "date_send", 25),
        ("Status", "status", 25),
        ("Resposta", "answer", 25),
        ("Data Resposta", "date_answer", 25),
        ("Especialidade", "type", 75)
    ]

    # Header
    header_fill = PatternFill("solid", fgColor="FFCD6B23")
    header_font = Font(color="FFFFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    last_id = None
    color_idx = 0

    for row_idx, item in enumerate(data, start=2):
        if last_id is None:
            last_id = item.get("id")

        if last_id != item.get("id"):
            color_idx = 1 if color_idx == 0 else 0

        fill = PatternFill("solid", fgColor=COLORS[color_idx])
        align = Alignment(horizontal="left", vertical="center")

        values = {
            "name": item.get("nome", " - "),
            "number": item.get("numero", " - "),
            "date": (
                f"{item.get('data_consulta').strftime('%d/%m/%Y')} {item.get('hora_consulta')}"
                if item.get("data_consulta") else " - "
            ),
            "date_send": (
                item.get("dt_envio").strftime("%d/%m/%Y %H:%M:%S")
                if item.get("dt_envio") else " - "
            ),
            "status": "Erro no Envio" if item.get("status_envio") != "S" else "Respondido" if item.get("resposta") is not None else "Pendente",
            "answer": item.get("resposta", " - "), 
            "date_answer": (
                item.get("dt_resposta").strftime("%d/%m/%Y %H:%M:%S")
                if item.get("dt_resposta") else " - "
            ),
            "type": item.get("especialidade", " - "),
        }

        for col_idx, (_, key, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values.get(key))
            cell.fill = fill
            cell.alignment = align

            if key in ("cpf", "phone"):
                cell.number_format = "@"

            if key == "transport_date" and isinstance(values[key], datetime):
                cell.number_format = "DD/MM/YYYY"

        ws.row_dimensions[row_idx].height = 20
        last_id = item.get("id")

    # Borders
    thin = Side(style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(columns)):
        for cell in row:
            cell.border = border

    # AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
